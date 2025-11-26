import os
import io
import json
import logging
from datetime import datetime, timezone
import uuid
import openpyxl
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_from_directory, Response, g, send_file
from flask_session import Session
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect, generate_csrf, CSRFError
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from openai import OpenAI
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from dotenv import load_dotenv
import secrets

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)

# Generate a secure secret key if not in environment variables
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)

# Configure application
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_SECRET_KEY'] = os.environ.get('CSRF_SECRET_KEY') or secrets.token_hex(32)
app.config['WTF_CSRF_TIME_LIMIT'] = 3600  # 1 hour CSRF token expiration

# Configure session to use filesystem instead of cookie-based
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'flask_session')
app.config['SESSION_FILE_THRESHOLD'] = 100  # Maximum number of sessions stored
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour session lifetime

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Ensure session directory exists
os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)

# Initialize extensions
db = SQLAlchemy(app)
csrf = CSRFProtect(app)

# Initialize Flask-Session
Session(app)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[logging.StreamHandler(), logging.FileHandler('app.log')])
logger = logging.getLogger(__name__)

# Initialize OpenAI client with API key from environment variable
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Create all database tables
with app.app_context():
    db.create_all()

# Add current datetime to template context
@app.context_processor
def inject_now():
    return {'now': datetime.now(timezone.utc)}

# Models
class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    bank_name = db.Column(db.String(100))
    account_number = db.Column(db.String(50))
    ifsc_code = db.Column(db.String(20))

class ReimbursementForm(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), db.ForeignKey('employee.employee_id', ondelete='SET NULL'), nullable=True)
    designation = db.Column(db.String(100))
    location = db.Column(db.String(100))
    from_date = db.Column(db.Date)
    to_date = db.Column(db.Date)
    total_amount = db.Column(db.Float)
    image_filename = db.Column(db.String(200))
    raw_data = db.Column(db.Text) # <-- Add this line
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

class ExpenseEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    form_id = db.Column(db.Integer, db.ForeignKey('reimbursement_form.id'), nullable=False)
    date = db.Column(db.Date)
    from_location = db.Column(db.String(100))
    to_location = db.Column(db.String(100))
    purpose = db.Column(db.String(200))
    mode_of_travel = db.Column(db.String(50))
    distance_km = db.Column(db.Float)
    amount_rs = db.Column(db.Float)

# Create database tables
with app.app_context():
    db.create_all()

# Add CSRF token to all templates
@app.context_processor
def inject_template_vars():
    return {
        'csrf_token': generate_csrf,
        'debug': app.debug
    }

# Handle CSRF errors
@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    logger.warning(f'CSRF Error: {str(e)}')
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': False, 
            'message': 'The form has expired. Please refresh the page and try again.',
            'requires_refresh': True
        }), 403
    return render_template('error.html', 
                         error='The form has expired. Please refresh the page and try again.',
                         status_code=403), 403

def clean_amount(amount_str):
    """Extract numeric value from amount string (e.g., 'Rs. 1011 only' -> 1011.0)"""
    if not amount_str:
        return 0.0
    
    # Remove non-numeric characters except decimal point
    import re
    cleaned = re.sub(r'[^\d.]', '', str(amount_str))
    try:
        return float(cleaned) if cleaned else 0.0
    except (ValueError, TypeError):
        return 0.0

# Helper function to parse dates
def parse_date(date_str):
    if not date_str or not isinstance(date_str, str):
        return None
        
    # Remove any whitespace and normalize date separators
    date_str = date_str.strip().replace('/', '.').replace('-', '.')
    
    # List of possible date formats to try
    date_formats = [
        '%d.%m.%Y',  # 01.01.2023
        '%Y.%m.%d',  # 2023.01.01
        '%d-%m-%Y',  # 01-01-2023
        '%Y-%m-%d',  # 2023-01-01
        '%d/%m/%Y',  # 01/01/2023
        '%m/%d/%Y',  # 01/01/2023 (US format)
        '%Y/%m/%d'   # 2023/01/01
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
            
    # If all formats fail, try to extract just the date part from datetime strings
    if 'T' in date_str:
        date_part = date_str.split('T')[0]
        try:
            return datetime.strptime(date_part, '%Y-%m-%d').date()
        except ValueError:
            pass
            
    logger.warning(f"Failed to parse date: {date_str}")
    return None

# Extract data using OpenAI API
def extract_data_with_openai(file_path):
    try:
        # Upload the image
        with open(file_path, 'rb') as f:
            file = client.files.create(
                file=f,
                purpose="user_data"
            )
        
        # Define the prompt with explicit date format and strict JSON response requirement
        prompt = (
            """
            Carefully examine the expense table on this form. Internally determine the layout including the number of rows and columns to ensure no values are skipped during extraction — especially merged cells or multi-line text.

            Then extract:

            1. Header Section (top of the form):
            - Employee ID, Name of Employee, Designation, Location
            - From Date and To Date (format: DD.MM.YYYY)
            - Total Amount (numeric only, no symbols).  
                It may be labeled as "Total" or "Total Amount" or "Total Amount (in Rs.)"  
                If this field is **not present anywhere in the image**, return 0.
            - In addition to the above, always compute and return the sum of all extracted "Amount (in Rs.)" values as a separate field called "Calculated Total".
            - If both a header total and a calculated total are present, compare them and add a boolean field "total_mismatch": true if they differ, false if they match. Always include both totals and the flag in the output.

            2. Expense Entries (the table):
            - Merged-cell values (e.g. Purpose, Date, From/To) must be duplicated across the rows they visually span.
            - For each row, extract:
                • Date (in DD.MM.YYYY format)
                • From, To
                • Purpose (exact text)
                • Mode of Travel: normalize to one of the following values:
                - "2-Wheeler"
                - "4-Wheeler"
                - "Cab" (for entries mentioning cab, ola, uber, auto, or taxi)
                - "Food & Misc." (see below)
                • Distance (in Km): extract numeric only, or zero
                • Amount (in Rs.): extract numeric only

            Special case:
            - For any row related to food or miscellaneous expenses:
                • Set **From**, **To**, and **Mode of Travel** to `"Food & Misc."`
                • Set **Distance (in Km)** to `"0"`
                • Purpose should remain as-is

            3. Validation:
            - Sum all extracted "Amount (in Rs.)" values.
            - If a Total Amount is present in the header, compare it to the computed sum.
            - If the header total is missing, simply use the computed sum as the total.

            Return ONLY a single JSON object — no markdown, no explanation, no comments:

            {
            "header": {
                "Employee ID": "string",
                "Name of Employee": "string",
                "Designation": "string",
                "Location": "string",
                "From Date": "DD.MM.YYYY",
                "To Date": "DD.MM.YYYY",
                "Total Amount": "string",
                "Calculated Total": "string",
                "total_mismatch": true/false
            },
            "expenses": [
                {
                "Date": "DD.MM.YYYY",
                "From": "string",
                "To": "string",
                "Purpose": "string",
                "Mode of Travel": "string",
                "Distance (in Km)": "string",
                "Amount (in Rs.)": "string"
                }
            ]
            }
            """
        )     
        # Call the model with the required input format
        response = client.responses.create(
            model="gpt-4.1",
            input=[
                {
                    "role": "user",
                    "content": [
                        {"type": "input_image", "file_id": file.id},
                        {"type": "input_text", "text": prompt}
                    ]
                }
            ]
        )
        
        # Extract the content from the response
        content = response.output[0].content[0].text
        
        # Clean the response (remove markdown code blocks if present)
        if content.strip().startswith('```'):
            # Handle both ```json and ``` cases
            if content.strip().startswith('```json'):
                content = content.split('```json')[1].split('```')[0].strip()
            else:
                content = content.split('```')[1].strip()
        
        if not content:
            logger.error("Empty response from OpenAI API")
            return None
        
        try:
            # Try to parse the JSON
            data = json.loads(content)
            
            # Validate the response structure
            if not isinstance(data, dict) or 'header' not in data or 'expenses' not in data:
                logger.error(f"Invalid response format. Missing required fields. Content: {content}")
                return None

            # Postprocess: If Total Amount is '0' or empty, replace with sum of expenses
            header = data.get('header', {})
            expenses = data.get('expenses', [])
            total_amount = header.get('Total Amount', '').strip()
            # Compute sum of all extracted expense amounts
            try:
                sum_expenses = sum(float(e.get('Amount (in Rs.)', '0').replace(',', '').strip()) for e in expenses)
            except Exception as e:
                logger.error(f"Error calculating sum of expenses: {e}")
                sum_expenses = 0

            if not total_amount or total_amount == '0':
                header['Total Amount'] = str(int(sum_expenses) if sum_expenses == int(sum_expenses) else sum_expenses)
            # Always update Calculated Total to backend sum for consistency
            header['Calculated Total'] = str(int(sum_expenses) if sum_expenses == int(sum_expenses) else sum_expenses)
            # Update mismatch flag
            try:
                total_amount_val = float(header['Total Amount'])
                calculated_total_val = float(header['Calculated Total'])
                header['total_mismatch'] = abs(total_amount_val - calculated_total_val) > 0.01
            except Exception as e:
                logger.warning(f"Could not compare totals for mismatch: {e}")
                header['total_mismatch'] = False

            logger.info(f"Successfully parsed data from OpenAI: {json.dumps(data, indent=2)}")
            return data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse response as JSON: {e}. Raw response: {content}")
            return None
            
    except Exception as e:
        logger.error(f"OpenAI API error: {str(e)}", exc_info=True)
        raise

# Routes
@app.route('/')
def index():
    # Get current date for the dashboard
    now = datetime.now()
    current_month = now.month
    current_year = now.year
    
    # Get counts for dashboard cards
    total_employees = Employee.query.count()
    total_forms = ReimbursementForm.query.count()
    
    # Get recent forms (last 5) with employee details
    recent_forms = ReimbursementForm.query.order_by(
        ReimbursementForm.created_at.desc()
    ).limit(5).all()
    
    # Build processed_recent_forms with robust employee name fallback
    processed_recent_forms = []
    for form in recent_forms:
        emp = Employee.query.filter_by(employee_id=form.employee_id).first()
        raw_data = None
        extracted_name = None
        if hasattr(form, 'raw_data') and form.raw_data:
            try:
                raw_data = form.raw_data if isinstance(form.raw_data, dict) else json.loads(form.raw_data)
                extracted_name = raw_data.get('header', {}).get('Name of Employee') if raw_data else None
            except Exception:
                raw_data = None
                extracted_name = None
        processed_recent_forms.append({
            'form': form,
            'employee': emp,
            'extracted_name': extracted_name
        })
    recent_forms = processed_recent_forms
    
    # Get monthly summary for the current month
    monthly_summary = db.session.query(
        func.sum(ExpenseEntry.amount_rs).label('total_amount'),
        func.count(ReimbursementForm.id.distinct()).label('form_count'),
        func.count(Employee.id.distinct()).label('employee_count')
    ).join(
        ReimbursementForm, ReimbursementForm.id == ExpenseEntry.form_id
    ).join(
        Employee, Employee.employee_id == ReimbursementForm.employee_id
    ).filter(
        func.strftime('%Y', ReimbursementForm.to_date) == str(current_year),
        func.strftime('%m', ReimbursementForm.to_date) == f"{current_month:02d}"
    ).first()
    
    # Get recent employees (last 5)
    recent_employees = Employee.query.order_by(
        Employee.id.desc()
    ).limit(5).all()
    
    return render_template(
        'index.html',
        total_employees=total_employees,
        total_forms=total_forms,
        recent_forms=processed_recent_forms,
        monthly_summary=monthly_summary,
        recent_employees=recent_employees,
        current_month=now.strftime('%B %Y'),
        now=now
    )

@app.route('/employees')
def employees():
    employees = Employee.query.all()
    return render_template('employees.html', employees=employees)

@app.route('/new_employee', methods=['GET', 'POST'])
def new_employee():
    if request.method == 'POST':
        employee_id = request.form['employee_id'].replace(' ', '')
        name = request.form['name']
        bank_name = request.form['bank_name']
        account_number = request.form['account_number']
        ifsc_code = request.form['ifsc_code']
        
        employee = Employee(employee_id=employee_id, name=name, bank_name=bank_name,
                            account_number=account_number, ifsc_code=ifsc_code)
        try:
            db.session.add(employee)
            db.session.commit()
            flash('Employee added successfully!')
        except IntegrityError:
            db.session.rollback()
            flash('Employee ID already exists.')
        return redirect(url_for('employees'))
    return render_template('new_employee.html')

@app.route('/edit_employee/<int:employee_id>', methods=['GET', 'POST'])
def edit_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    if request.method == 'POST':
        employee_id_clean = request.form['employee_id'].replace(' ', '')
        employee.employee_id = employee_id_clean
        employee.name = request.form['name']
        employee.bank_name = request.form['bank_name']
        employee.account_number = request.form['account_number']
        employee.ifsc_code = request.form['ifsc_code']
        db.session.commit()
        flash('Employee updated successfully!', 'success')
        return redirect(url_for('employees'))
    return render_template('edit_employee.html', employee=employee)

@app.route('/run_job', methods=['GET', 'POST'])
def run_job():
    if request.method == 'POST':
        files = request.files.getlist('images')
        # Ensure upload folder exists
        upload_folder = app.config['UPLOAD_FOLDER']
        if not os.path.exists(upload_folder):
            try:
                os.makedirs(upload_folder, exist_ok=True)
                logger.info(f"Created upload folder: {os.path.abspath(upload_folder)}")
            except Exception as e:
                logger.error(f"Failed to create upload folder: {upload_folder}: {e}")
                flash(f"Server error: Could not create upload folder.", 'error')
                return redirect(url_for('run_job'))

        from pathlib import Path
        from image_preprocess import autocrop_image
        for file in files:
            orig_filename = file.filename
            ext = os.path.splitext(orig_filename)[1]
            unique_filename = f"{uuid.uuid4().hex}{ext}"
            filepath = os.path.join(upload_folder, unique_filename)
            try:
                file.save(filepath)
            except Exception as e:
                logger.error(f"Failed to save uploaded file {orig_filename} to {filepath}: {e}")
                flash(f"Failed to save uploaded file {orig_filename}.", 'error')
                continue
            # Check file existence and size
            if not os.path.exists(filepath):
                logger.error(f"File {filepath} does not exist after save!")
                flash(f"File {orig_filename} could not be found after upload.", 'error')
                continue
            file_size = os.path.getsize(filepath)
            logger.info(f"Saved file {orig_filename} as {unique_filename} to {filepath} (size: {file_size} bytes)")
            if file_size == 0:
                logger.error(f"File {filepath} is empty after save!")
                flash(f"Uploaded file {orig_filename} is empty after save. Please try again.", 'error')
                continue
            # Preprocess the image (autocrop, enhance, combine)
            try:
                processed_path = autocrop_image(filepath, Path(upload_folder))
                processed_filename = os.path.basename(processed_path)
                logger.info(f"Preprocessed image saved as {processed_filename}")
            except Exception as e:
                logger.error(f"Error preprocessing image {filepath}: {e}")
                flash(f"Failed to preprocess image {orig_filename}. Error: {e}", 'error')
                continue
            # Extract data using OpenAI on the preprocessed image
            data = extract_data_with_openai(str(processed_path))
            if data is None:
                flash(f"Failed to extract data from {orig_filename}")
                continue
            
            header = data['header']
            expenses = data['expenses']
            
            # Parse dates with detailed error handling
            from_date_str = header.get('From Date', '')
            to_date_str = header.get('To Date', '')
            
            from_date = parse_date(from_date_str)
            to_date = parse_date(to_date_str)
            
            if not from_date:
                flash(f"Invalid 'From Date' format: '{from_date_str}'. Expected format: DD.MM.YYYY, YYYY-MM-DD, or similar", 'error')
                logger.error(f"Failed to parse 'From Date': {from_date_str} in file: {filename}")
                continue
                
            if not to_date:
                flash(f"Invalid 'To Date' format: '{to_date_str}'. Expected format: DD.MM.YYYY, YYYY-MM-DD, or similar", 'error')
                logger.error(f"Failed to parse 'To Date': {to_date_str} in file: {filename}")
                continue
                
            if to_date < from_date:
                flash(f"'To Date' ({to_date}) cannot be before 'From Date' ({from_date}) in file: {filename}", 'error')
                continue
            
            # Create ReimbursementForm
            form = ReimbursementForm(
                employee_id=header.get('Employee ID', '').replace(' ', ''),
                designation=header.get('Designation', ''),
                location=header.get('Location', ''),
                from_date=from_date,
                to_date=to_date,
                total_amount=clean_amount(header.get('Total Amount', '0')),
                image_filename=processed_filename,
                raw_data=json.dumps(data)  # <-- Save the extracted data!
            )
            try:
                db.session.add(form)
                db.session.flush()
                
                # Create ExpenseEntry records
                for exp in expenses:
                    entry_date_str = exp.get('Date', '')
                    entry_date = parse_date(entry_date_str) if entry_date_str else None
                    if not entry_date:
                        continue
                    distance = clean_amount(exp.get('Distance (in Km)', '0'))
                    amount = clean_amount(exp.get('Amount (in Rs.)', '0'))
                    entry = ExpenseEntry(
                        form_id=form.id,
                        date=entry_date,
                        from_location=exp.get('From', ''),
                        to_location=exp.get('To', ''),
                        purpose=exp.get('Purpose', ''),
                        mode_of_travel=exp.get('Mode of Travel', ''),
                        distance_km=distance,
                        amount_rs=amount
                    )
                    db.session.add(entry)
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash(f"Employee ID {header.get('Employee ID', '')} does not exist for {filename}")
            except Exception as e:
                db.session.rollback()
                flash(f"Error processing {filename}: {str(e)}")
        
        flash('Job completed successfully!')
        return redirect(url_for('index'))
    return render_template('run_job.html')

@app.route('/monthly_summary_selector')
def monthly_summary_selector():
    """Show a form to select month and year for summary"""
    # Get current date
    now = datetime.now()
    current_year = now.year
    current_month = now.month
    
    # Get distinct months with data for the recent months list
    recent_months = db.session.query(
        func.strftime('%Y', ReimbursementForm.to_date).label('year'),
        func.strftime('%m', ReimbursementForm.to_date).cast(db.Integer).label('month')
    ).distinct().order_by(
        func.strftime('%Y', ReimbursementForm.to_date).desc(),
        func.strftime('%m', ReimbursementForm.to_date).desc()
    ).limit(6).all()
    
    return render_template('monthly_summary_selector.html',
                         current_year=current_year,
                         current_month=current_month,
                         recent_months=recent_months,
                         month_names={
                             1: 'January', 2: 'February', 3: 'March', 4: 'April',
                             5: 'May', 6: 'June', 7: 'July', 8: 'August',
                             9: 'September', 10: 'October', 11: 'November', 12: 'December'
                         })

@app.route('/monthly_summary')
def monthly_summary():
    """Show monthly summary for the selected month and year"""
    try:
        year = int(request.args.get('year', datetime.now().year))
        month = int(request.args.get('month', datetime.now().month))
    except (ValueError, TypeError):
        flash('Invalid year or month specified', 'error')
        return redirect(url_for('monthly_summary_selector'))
    
    return redirect(url_for('monthly_summary_detail', year=year, month=month))

@app.route('/monthly_summary/<int:year>/<int:month>')
def monthly_summary_detail(year, month):
    # Only these modes are allowed for local conveyance
    local_modes = {'2-wheeler', '4-wheeler', 'cab'}
    food_misc_mode = 'food & misc.'

    # Get month name for display
    month_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    month_name = month_names[month - 1] if 1 <= month <= 12 else f'Month {month}'

    # Query for forms ending in the specified month
    forms = ReimbursementForm.query.filter(
        func.strftime('%Y', ReimbursementForm.to_date) == str(year),
        func.strftime('%m', ReimbursementForm.to_date) == f'{month:02d}'
    ).all()

    summary = {}
    for form in forms:
        emp = Employee.query.filter_by(employee_id=form.employee_id).first() if form.employee_id else None
        # Fallback: Try to get extracted name from raw_data if no DB employee
        extracted_name = None
        if not emp and getattr(form, 'raw_data', None):
            try:
                import json
                raw = form.raw_data if isinstance(form.raw_data, dict) else json.loads(form.raw_data)
                extracted_name = raw.get('header', {}).get('Name of Employee')
            except Exception:
                extracted_name = None
        emp_key = form.employee_id or f'form_{form.id}'
        if emp_key not in summary:
            summary[emp_key] = {
                'employee_id': form.employee_id or '',
                'name': emp.name if emp and emp.name else (extracted_name or ''),
                'location': form.location or '',
                'exp_reimbursement': 0,
                'local_conveyance': 0,
                'bank_name': emp.bank_name if emp else '',
                'account_number': emp.account_number if emp else '',
                'ifsc_code': emp.ifsc_code if emp else ''
            }
        entries = ExpenseEntry.query.filter_by(form_id=form.id).all()
        for entry in entries:
            # Normalize mode for strict matching
            mode = (entry.mode_of_travel or '').strip().lower().replace('-', '').replace('&', 'and').replace(' ', '')
            # Map allowed modes to normalized forms
            allowed_modes = {
                '2-wheeler': '2wheeler',
                '4-wheeler': '4wheeler',
                'cab': 'cab',
                'food & misc.': 'foodandmisc.'
            }
            if mode == allowed_modes['2-wheeler'] or mode == allowed_modes['4-wheeler'] or mode == allowed_modes['cab']:
                summary[emp_key]['local_conveyance'] += (entry.amount_rs or 0)
            else:
                summary[emp_key]['exp_reimbursement'] += (entry.amount_rs or 0)

    summary_list = sorted(
        [
            {
                'employee_id': data['employee_id'],
                'name': data['name'],
                'location': data['location'],
                'exp_reimbursement': data['exp_reimbursement'],
                'local_conveyance': data['local_conveyance'],
                'total_payable': data['exp_reimbursement'] + data['local_conveyance'],
                'bank_name': data['bank_name'],
                'account_number': data['account_number'],
                'ifsc_code': data['ifsc_code']
            }
            for emp_key, data in summary.items()
        ],
        key=lambda x: x['name'].lower() if x['name'] else x['employee_id']
    )

    totals = {
        'reimbursement': sum(item['exp_reimbursement'] for item in summary_list),
        'local': sum(item['local_conveyance'] for item in summary_list),
        'grand_total': sum(item['total_payable'] for item in summary_list)
    }

    return render_template('monthly_summary.html',
                         summary=summary_list,
                         year=year,
                         month=month,
                         month_name=month_name,
                         totals=totals,
                         month_names=month_names)

@app.route('/forms')
def view_forms():
    """View all processed reimbursement forms with filtering options"""
    # Get filter parameters
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    employee_id = request.args.get('employee_id', '').strip()
    
    # Start with base query
    query = ReimbursementForm.query
    
    # Apply filters
    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            query = query.filter(ReimbursementForm.from_date >= from_date)
        except ValueError:
            flash('Invalid from date format. Use YYYY-MM-DD.', 'error')
    
    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
            query = query.filter(ReimbursementForm.to_date <= to_date)
        except ValueError:
            flash('Invalid to date format. Use YYYY-MM-DD.', 'error')
    
    if employee_id:
        query = query.filter(ReimbursementForm.employee_id.ilike(f'%{employee_id}%'))
    
    # Order by most recent first
    forms = query.order_by(ReimbursementForm.to_date.desc()).all()
    
    # Get employee details for display
    form_data = []
    for form in forms:
        emp = Employee.query.filter_by(employee_id=form.employee_id).first()
        raw_data = getattr(form, 'raw_data', None)
        extracted_name = None
        if raw_data:
            try:
                if isinstance(raw_data, str):
                    raw_data = json.loads(raw_data)
                extracted_name = raw_data.get('header', {}).get('Name of Employee') if raw_data else None
            except Exception:
                extracted_name = None
        form_data.append({
            'id': form.id,
            'employee_id': form.employee_id,
            'designation': form.designation or '',
            'location': form.location or '',
            'from_date': form.from_date,
            'to_date': form.to_date,
            'total_amount': form.total_amount or 0,
            'image_filename': form.image_filename,
            'employee': emp,
            'extracted_name': extracted_name
        })
    
    return render_template('view_forms.html', forms=form_data)

@app.route('/form_details/<int:form_id>')
def form_details(form_id):
    """Get details of a specific form (for AJAX loading)"""
    try:
        logger.info(f"Fetching details for form ID: {form_id}")
        form = ReimbursementForm.query.get(form_id)
        if not form:
            logger.error(f"Form with ID {form_id} not found")
            return "Form not found", 404
            
        logger.debug(f"Found form: {form.id}, Employee ID: {form.employee_id}")
        
        entries = ExpenseEntry.query.filter_by(form_id=form_id).order_by(ExpenseEntry.date).all()
        emp = Employee.query.filter_by(employee_id=form.employee_id).first()
        raw_data = None
        extracted_name = None
        if hasattr(form, 'raw_data') and form.raw_data:
            try:
                raw_data = form.raw_data if isinstance(form.raw_data, dict) else json.loads(form.raw_data)
                extracted_name = raw_data.get('header', {}).get('Name of Employee') if raw_data else None
            except Exception:
                raw_data = None
                extracted_name = None
        return render_template('_form_details.html',
                            form=form,
                            entries=entries,
                            employee=emp,
                            extracted_name=extracted_name)
    except Exception as e:
        logger.error(f"Error in form_details route: {str(e)}", exc_info=True)
        if app.debug:
            import traceback
            return f"An error occurred: {str(e)}<br><pre>{traceback.format_exc()}</pre>", 500
        return "An error occurred while loading form details. Please try again later.", 500

@app.route('/form_details_ajax/<int:form_id>')
def form_details_ajax(form_id):
    """Get details of a specific form (for AJAX loading)"""
    try:
        logger.info(f"[AJAX] Fetching details for form ID: {form_id}")
        form = ReimbursementForm.query.get(form_id)
        if not form:
            logger.error(f"[AJAX] Form with ID {form_id} not found")
            return "Form not found", 404
            
        logger.debug(f"[AJAX] Found form: {form.id}, Employee ID: {form.employee_id}")
        
        entries = ExpenseEntry.query.filter_by(form_id=form_id).order_by(ExpenseEntry.date).all()
        emp = Employee.query.filter_by(employee_id=form.employee_id).first()
        raw_data = None
        extracted_name = None
        if hasattr(form, 'raw_data') and form.raw_data:
            try:
                raw_data = form.raw_data if isinstance(form.raw_data, dict) else json.loads(form.raw_data)
                extracted_name = raw_data.get('header', {}).get('Name of Employee') if raw_data else None
            except Exception:
                raw_data = None
                extracted_name = None
        return render_template('_form_details_ajax.html',
                            form=form,
                            entries=entries,
                            employee=emp,
                            extracted_name=extracted_name)
    except Exception as e:
        logger.error(f"[AJAX] Error in form_details_ajax route: {str(e)}", exc_info=True)
        if app.debug:
            import traceback
            return f"An error occurred: {str(e)}<br><pre>{traceback.format_exc()}</pre>", 500
        return "An error occurred while loading form details. Please try again later.", 500

@app.route('/admin/edit_form/<int:form_id>', methods=['GET', 'POST'])
def edit_form(form_id):
    """Edit form data including header and expense entries"""
    form_data = ReimbursementForm.query.get_or_404(form_id)
    employee = Employee.query.filter_by(employee_id=form_data.employee_id).first()
    entries = ExpenseEntry.query.filter_by(form_id=form_id).order_by(ExpenseEntry.date).all()

    extracted_name = None
    if form_data.raw_data:
        try:
            import json
            raw = form_data.raw_data if isinstance(form_data.raw_data, dict) else json.loads(form_data.raw_data)
            extracted_name = raw.get('header', {}).get('Name of Employee')
        except Exception:
            extracted_name = None

    if request.method == 'POST':
        try:
            # Update form header data
            form_data.employee_id = request.form.get('employee_id', form_data.employee_id).replace(' ', '')
            form_data.designation = request.form.get('designation', form_data.designation)
            form_data.location = request.form.get('location', form_data.location)
            form_data.from_date = datetime.strptime(request.form['from_date'], '%Y-%m-%d')
            form_data.to_date = datetime.strptime(request.form['to_date'], '%Y-%m-%d')
            form_data.total_amount = float(request.form.get('total_amount', 0))
            
            # Update employee details if they exist
            if employee:
                employee.name = request.form.get('employee_name', employee.name)
                employee.bank_name = request.form.get('bank_name', employee.bank_name)
                employee.account_number = request.form.get('account_number', employee.account_number)
                employee.ifsc_code = request.form.get('ifsc_code', employee.ifsc_code)
            
            # Update or create expense entries
            entry_ids = []
            for i, entry in enumerate(request.form.getlist('entry_id')):
                entry_data = {
                    'id': int(entry) if entry else None,
                    'date': datetime.strptime(request.form.getlist('entry_date')[i], '%Y-%m-%d'),
                    'from_location': request.form.getlist('entry_from')[i],
                    'to_location': request.form.getlist('entry_to')[i],
                    'purpose': request.form.getlist('entry_purpose')[i],
                    'mode_of_travel': request.form.getlist('entry_mode')[i],
                    'distance_km': float(request.form.getlist('entry_distance')[i] or 0),
                    'amount_rs': float(request.form.getlist('entry_amount')[i] or 0)
                }
                
                if entry_data['id']:
                    # Update existing entry
                    db_entry = db.session.get(ExpenseEntry, entry_data['id'])
                    if db_entry:
                        for key, value in entry_data.items():
                            if key != 'id' and hasattr(db_entry, key):
                                setattr(db_entry, key, value)
                        entry_ids.append(db_entry.id)
                else:
                    # Create new entry
                    new_entry = ExpenseEntry(
                        form_id=form_id,
                        **{k: v for k, v in entry_data.items() if k != 'id'}
                    )
                    db.session.add(new_entry)
                    db.session.flush()
                    entry_ids.append(new_entry.id)
            
            # Delete entries that were removed
            ExpenseEntry.query.filter(
                ExpenseEntry.form_id == form_id,
                ~ExpenseEntry.id.in_(entry_ids)
            ).delete(synchronize_session=False)
            
            db.session.commit()
            flash('Form updated successfully!', 'success')
            return redirect(url_for('view_forms'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating form: {str(e)}")
            flash('Error updating form. Please try again.', 'error')
    
    return render_template('admin/edit_form.html',
                         form=form_data,
                         entries=entries,
                         employee=employee,
                         extracted_name=extracted_name)

@app.route('/delete_form/<int:form_id>', methods=['POST'])
@csrf.exempt  # Temporarily disable CSRF for this endpoint to test
# @login_required  # Uncomment if you have authentication
# @admin_required  # Uncomment if you have admin requirements
def delete_form(form_id):
    """Delete a form and its entries"""
    # Get form data
    form_data = request.form
    
    # For debugging - log the incoming request
    app.logger.info(f"Delete form request for ID: {form_id}")
    app.logger.info(f"Form data: {dict(form_data)}")
    app.logger.info(f"Headers: {dict(request.headers)}")
    
    # Get the form
    form = ReimbursementForm.query.get_or_404(form_id)
    
    try:
        # Delete associated expense entries
        ExpenseEntry.query.filter_by(form_id=form_id).delete()
        
        # Delete the form
        db.session.delete(form)
        db.session.commit()
        
        success_message = 'Form and associated entries deleted successfully!'
        app.logger.info(success_message)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': 'Form deleted successfully',
                'redirect': url_for('view_forms')
            })
            
        return redirect(url_for('view_forms'))
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting form {form_id}: {str(e)}")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False, 
                'message': f'Error deleting form: {str(e)}',
                'error': str(e)
            }), 500
            
        flash('Error deleting form. Please try again.', 'error')
        return redirect(url_for('view_forms'))

# --- BULK DELETE ROUTE ---
@app.route('/bulk_delete_forms', methods=['POST'])
@csrf.exempt  # Remove this after CSRF AJAX is confirmed working
# @login_required
# @admin_required
def bulk_delete_forms():
    """Bulk delete reimbursement forms and their expense entries"""
    try:
        data = request.get_json()
        form_ids = data.get('form_ids', [])
        if not form_ids:
            return jsonify({'success': False, 'message': 'No forms selected.'}), 400
        for form_id in form_ids:
            form = ReimbursementForm.query.get(form_id)
            if form:
                ExpenseEntry.query.filter_by(form_id=form.id).delete()
                db.session.delete(form)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Selected forms deleted.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/export_monthly_summary', methods=['POST'])
def export_monthly_summary():
    try:
        year = int(request.form.get('year', datetime.now().year))
        month = int(request.form.get('month', datetime.now().month))

        # Only these modes are allowed for local conveyance
        local_modes = {'2-wheeler', '4-wheeler', 'cab'}
        food_misc_mode = 'food & misc.'

        # Query for forms ending in the specified month
        forms = ReimbursementForm.query.filter(
            func.strftime('%Y', ReimbursementForm.to_date) == str(year),
            func.strftime('%m', ReimbursementForm.to_date) == f'{month:02d}'
        ).all()

        summary = {}
        for form in forms:
            emp = Employee.query.filter_by(employee_id=form.employee_id).first() if form.employee_id else None
            extracted_name = None
            if not emp and getattr(form, 'raw_data', None):
                try:
                    import json
                    raw = form.raw_data if isinstance(form.raw_data, dict) else json.loads(form.raw_data)
                    extracted_name = raw.get('header', {}).get('Name of Employee')
                except Exception:
                    extracted_name = None
            emp_key = form.employee_id or f'form_{form.id}'
            if emp_key not in summary:
                summary[emp_key] = {
                    'employee_id': form.employee_id or '',
                    'name': emp.name if emp and emp.name else (extracted_name or ''),
                    'location': form.location or '',
                    'exp_reimbursement': 0,
                    'local_conveyance': 0,
                    'bank_name': emp.bank_name if emp else '',
                    'account_number': emp.account_number if emp else '',
                    'ifsc_code': emp.ifsc_code if emp else ''
                }
            entries = ExpenseEntry.query.filter_by(form_id=form.id).all()
            for entry in entries:
                mode = (entry.mode_of_travel or '').strip().lower().replace('-', '').replace('&', 'and').replace(' ', '')
                allowed_modes = {
                    '2-wheeler': '2wheeler',
                    '4-wheeler': '4wheeler',
                    'cab': 'cab',
                    'food & misc.': 'foodandmisc.'
                }
                if mode == allowed_modes['2-wheeler'] or mode == allowed_modes['4-wheeler'] or mode == allowed_modes['cab']:
                    summary[emp_key]['local_conveyance'] += (entry.amount_rs or 0)
                else:
                    summary[emp_key]['exp_reimbursement'] += (entry.amount_rs or 0)

        summary_list = sorted(
            [
                {
                    'Employee ID': data['employee_id'],
                    'Employee Name': data['name'],
                    'Location': data['location'],
                    'Expense Reimbursement (₹)': data['exp_reimbursement'],
                    'Local Conveyance (₹)': data['local_conveyance'],
                    'Total Payable (₹)': data['exp_reimbursement'] + data['local_conveyance'],
                    'Bank Name': data['bank_name'],
                    'Account Number': data['account_number'],
                    'IFSC Code': data['ifsc_code']
                }
                for emp_key, data in summary.items()
            ],
            key=lambda x: x['Employee Name'].lower() if x['Employee Name'] else x['Employee ID']
        )

        if not summary_list:
            flash('No data available for the selected period', 'warning')
            return redirect(url_for('monthly_summary_selector'))

        # Create Excel file in memory
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{month:02d}-{year} Summary"

        # Add headers
        headers = list(summary_list[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DDDDDD")

        # Add data
        for row_num, row_data in enumerate(summary_list, 2):
            for col_num, key in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data[key])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 30)

        # Add totals row
        total_row = len(summary_list) + 3
        ws.cell(row=total_row, column=1, value="TOTALS:").font = openpyxl.styles.Font(bold=True)
        for col_num, key in enumerate(headers[3:6], 4):  # Only sum the numeric columns
            if col_num <= 6:  # Expense, Local, Total columns
                col_letter = openpyxl.utils.get_column_letter(col_num)
                ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}2:{col_letter}{len(summary_list) + 1})"
                ws[f"{col_letter}{total_row}"].number_format = '#,##0.00'
        for row in ws.iter_rows(min_row=2, max_row=len(summary_list) + 1, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = '#,##0.00'
        wb.save(output)
        output.seek(0)
        month_name = datetime(year, month, 1).strftime('%B')
        filename = f"Expense_Summary_{month_name}_{year}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error exporting monthly summary: {str(e)}")
        flash('Error generating Excel file. Please try again.', 'error')
        return redirect(url_for('monthly_summary_selector'))

@app.route('/export_form_excel/<int:form_id>', methods=['GET', 'POST'])
def export_form_excel(form_id):
    form = ReimbursementForm.query.get_or_404(form_id)
    entries = ExpenseEntry.query.filter_by(form_id=form_id).all()
    # Try to get extracted header from raw_data
    header = {}
    if getattr(form, 'raw_data', None):
        try:
            import json
            raw = form.raw_data if isinstance(form.raw_data, dict) else json.loads(form.raw_data)
            header = raw.get('header', {})
        except Exception:
            header = {}
    # Fallbacks
    header_fields = [
        ('Employee ID', form.employee_id or ''),
        ('Name of Employee', header.get('Name of Employee', '')),
        ('Designation', form.designation or header.get('Designation', '')),
        ('Location', form.location or header.get('Location', '')),
        ('From Date', form.from_date.strftime('%d-%m-%Y') if form.from_date else header.get('From Date', '')),
        ('To Date', form.to_date.strftime('%d-%m-%Y') if form.to_date else header.get('To Date', '')),
        ('Total Amount', form.total_amount or header.get('Total Amount', '')),
    ]
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Form_{form_id}"
    # Write header section
    row = 1
    for label, value in header_fields:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1
    # Write expense entries
    entry_headers = ['Date', 'From', 'To', 'Purpose', 'Mode of Travel', 'Distance (Km)', 'Amount (₹)']
    for col_num, h in enumerate(entry_headers, 1):
        ws.cell(row=row, column=col_num, value=h).font = openpyxl.styles.Font(bold=True)
    for entry in entries:
        row += 1
        ws.cell(row=row, column=1, value=entry.date.strftime('%d-%m-%Y') if entry.date else '')
        ws.cell(row=row, column=2, value=entry.from_location)
        ws.cell(row=row, column=3, value=entry.to_location)
        ws.cell(row=row, column=4, value=entry.purpose)
        ws.cell(row=row, column=5, value=entry.mode_of_travel)
        ws.cell(row=row, column=6, value=entry.distance_km)
        ws.cell(row=row, column=7, value=entry.amount_rs)
    wb.save(output)
    output.seek(0)
    filename = f"Form_{form_id}_Details.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/delete_employee/<int:employee_id>', methods=['POST'])
@csrf.exempt  # Remove this if you want CSRF protection and your form includes the token
def delete_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    db.session.delete(employee)
    db.session.commit()
    flash('Employee deleted successfully!', 'success')
    return redirect(url_for('employees'))

@app.route('/api/recent_forms')
def api_recent_forms():
    recent_forms = ReimbursementForm.query.order_by(ReimbursementForm.created_at.desc()).limit(5).all()
    result = []
    for form in recent_forms:
        emp = Employee.query.filter_by(employee_id=form.employee_id).first()
        raw_data = None
        extracted_name = None
        if hasattr(form, 'raw_data') and form.raw_data:
            try:
                raw_data = form.raw_data if isinstance(form.raw_data, dict) else json.loads(form.raw_data)
                extracted_name = raw_data.get('header', {}).get('Name of Employee') if raw_data else None
            except Exception:
                raw_data = None
                extracted_name = None
        result.append({
            'id': form.id,
            'employee_id': form.employee_id,
            'name': emp.name if emp and emp.name else (extracted_name or ''),
            'from_date': form.from_date.strftime('%d-%b-%Y') if form.from_date else '',
            'to_date': form.to_date.strftime('%d-%b-%Y') if form.to_date else '',
            'total_amount': form.total_amount or 0,
        })
    return jsonify(result)

@app.route('/api/forms')
def api_forms():
    forms = ReimbursementForm.query.order_by(ReimbursementForm.to_date.desc()).all()
    result = []
    for form in forms:
        emp = Employee.query.filter_by(employee_id=form.employee_id).first()
        raw_data = getattr(form, 'raw_data', None)
        extracted_name = None
        if raw_data:
            try:
                if isinstance(raw_data, str):
                    raw_data = json.loads(raw_data)
                extracted_name = raw_data.get('header', {}).get('Name of Employee') if raw_data else None
            except Exception:
                extracted_name = None
        result.append({
            'id': form.id,
            'employee_id': form.employee_id,
            'name': emp.name if emp and emp.name else (extracted_name or ''),
            'from_date': form.from_date.strftime('%d-%b-%Y') if form.from_date else '',
            'to_date': form.to_date.strftime('%d-%b-%Y') if form.to_date else '',
            'location': form.location or '',
            'total_amount': form.total_amount or 0,
        })
    return jsonify(result)

if __name__ == '__main__':
    app.run(debug=True, port=5001)